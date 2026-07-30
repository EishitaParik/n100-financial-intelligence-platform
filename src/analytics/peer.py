import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

class PeerComparisonEngine:

    def __init__(self):

        self.project_root = Path(__file__).resolve().parents[2]

        self.db_path = self.project_root / "nifty100.db"

        self.output_path = self.project_root / "output"
        self.output_path.mkdir(exist_ok=True)

        self.conn = sqlite3.connect(self.db_path)

    def load_data(self):

        query = """
        SELECT

            pg.peer_group_name,
            pg.is_benchmark,

            c.company_name,
            fr.company_id,
            fr.year,

            fr.return_on_equity_pct,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,
            fr.debt_to_equity,
            fr.interest_coverage,
            fr.asset_turnover,
            fr.free_cash_flow_cr,

            mc.market_cap_crore,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.dividend_yield_pct,

            a.compounded_sales_growth,
            a.compounded_profit_growth

        FROM peer_groups pg

        LEFT JOIN financial_ratios fr
            ON pg.company_id = fr.company_id

        LEFT JOIN companies c
            ON pg.company_id = c.id

        LEFT JOIN market_cap mc
            ON fr.company_id = mc.company_id
            AND fr.year = mc.year

        LEFT JOIN analysis a
            ON fr.company_id = a.company_id

        ORDER BY
            pg.peer_group_name,
            fr.company_id,
            fr.year
        """

        df = pd.read_sql(query, self.conn)

        numeric_columns = [

            "return_on_equity_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "asset_turnover",
            "free_cash_flow_cr",
            "market_cap_crore",
            "pe_ratio",
            "pb_ratio",
            "dividend_yield_pct",
            "compounded_sales_growth",
            "compounded_profit_growth"

        ]

        for col in numeric_columns:

            if col not in df.columns:
                continue

            df[col] = (
                df[col]
                .astype(str)
                .str.replace("%", "", regex=False)
                .str.replace(",", "", regex=False)
                .str.extract(r"(-?\d+\.?\d*)")[0]
            )

            df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

    def calculate_percentiles(self, df):

        metrics = [

            "return_on_equity_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "asset_turnover",
            "free_cash_flow_cr",
            "compounded_sales_growth",
            "compounded_profit_growth",
            "market_cap_crore",
            "pe_ratio",
            "debt_to_equity"

        ]

        result = df.copy()

        for metric in metrics:

            if metric not in result.columns:
                continue

            percentile_column = metric + "_percentile"

            result[percentile_column] = None

            for group in result["peer_group_name"].dropna().unique():

                mask = result["peer_group_name"] == group

                values = result.loc[mask, metric]

                if metric == "debt_to_equity":

                    percentile = (
                        values.rank(
                            pct=True,
                            ascending=False
                        ) * 100
                    )

                elif metric == "pe_ratio":

                    percentile = (
                        values.rank(
                            pct=True,
                            ascending=False
                        ) * 100
                    )

                else:

                    percentile = (
                        values.rank(
                            pct=True,
                            ascending=True
                        ) * 100
                    )

                result.loc[mask, percentile_column] = percentile.round(2)

        return result

    def export_excel(self, df):

        output_file = self.output_path / "peer_comparison.xlsx"

        green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
        red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        gold = PatternFill(fill_type="solid", start_color="FFD966", end_color="FFD966")

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            for group in sorted(df["peer_group_name"].dropna().unique()):

                sheet = group.replace("/", "-")[:31]

                group_df = (
                    df[df["peer_group_name"] == group]
                    .sort_values("company_name")
                    .copy()
                )

                # Median Row
                numeric_cols = group_df.select_dtypes(include="number").columns

                median_row = {}

                for col in group_df.columns:

                    if col == "company_name":
                        median_row[col] = "Peer Median"

                    elif col in numeric_cols:
                        median_row[col] = group_df[col].median()

                    else:
                        median_row[col] = ""

                group_df = pd.concat(
                    [group_df, pd.DataFrame([median_row])],
                    ignore_index=True
                )

                group_df.to_excel(
                    writer,
                    sheet_name=sheet,
                    index=False
                )

                worksheet = writer.sheets[sheet]

                headers = {
                    cell.value: cell.column
                    for cell in worksheet[1]
                }

                # Percentile colouring
                for header, col in headers.items():

                    if "percentile" not in str(header).lower():
                        continue

                    for row in range(2, worksheet.max_row):

                        cell = worksheet.cell(row=row, column=col)

                        if cell.value is None:
                                continue

                        try:
                                value = float(cell.value)
                        except (TypeError, ValueError):
                                continue

                        if value >= 75:
                                cell.fill = green

                        elif value >= 25:
                                cell.fill = yellow

                        else:
                                cell.fill = red   

                # Benchmark Highlight
                if "is_benchmark" in headers:

                    bench_col = headers["is_benchmark"]

                    for row in range(2, worksheet.max_row):

                        if worksheet.cell(row=row, column=bench_col).value:

                            for col in range(1, worksheet.max_column + 1):

                                worksheet.cell(row=row, column=col).fill = gold

        print("\nExcel exported successfully")
        print(output_file)

    def run(self):

        print("=" * 80)
        print("PEER COMPARISON ENGINE")
        print("=" * 80)

        df = self.load_data()

        print("Rows Loaded :", len(df))

        df = self.calculate_percentiles(df)

        self.export_excel(df)

        print("\nCompleted Successfully.")

    def close(self):
        self.conn.close()


if __name__ == "__main__":

    engine = PeerComparisonEngine()

    try:
        engine.run()

    finally:
        engine.close()