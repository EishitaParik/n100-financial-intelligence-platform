import re
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from openpyxl.styles import PatternFill


class ScreenerEngine:
    def __init__(self):
        self.project_root = Path(__file__).resolve().parents[2]
        self.db_path = self.project_root / "nifty100.db"
        self.output_path = self.project_root / "output"
        self.output_path.mkdir(exist_ok=True)
        self.conn = sqlite3.connect(self.db_path)

    def load_config(self):
        config_path = self.project_root / "config" / "screener_config.yaml"

        with open(config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    @staticmethod
    def extract_cagr(value, years):
        if pd.isna(value):
            return np.nan

        text = str(value)

        pattern = rf"{years}\s*Years?\s*:\s*(-?\d+(?:\.\d+)?)\s*%"
        match = re.search(pattern, text, re.IGNORECASE)

        if match:
            return float(match.group(1))

        return np.nan

    def load_data(self):
        query = """
        SELECT
            fr.company_id,
            fr.year,

            fr.return_on_equity_pct,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,
            fr.debt_to_equity,
            fr.interest_coverage,
            fr.asset_turnover,
            fr.free_cash_flow_cr,
            fr.capex_cr,
            fr.earnings_per_share,
            fr.book_value_per_share,
            fr.dividend_payout_ratio_pct,
            fr.total_debt_cr,
            fr.cash_from_operations_cr,

            c.company_name,
            c.roce_percentage,
            c.roe_percentage,

            pl.sales,
            pl.net_profit,
            pl.eps,
            pl.opm_percentage,

            a.compounded_sales_growth,
            a.compounded_profit_growth,
            a.stock_price_cagr,

            mc.market_cap_crore,
            mc.enterprise_value_crore,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.ev_ebitda,
            mc.dividend_yield_pct,

            s.broad_sector,
            s.sub_sector,
            s.market_cap_category

        FROM financial_ratios fr

        LEFT JOIN companies c
            ON fr.company_id = c.id

        LEFT JOIN profitandloss pl
            ON fr.company_id = pl.company_id
            AND fr.year = pl.year

        LEFT JOIN analysis a
            ON fr.company_id = a.company_id

        LEFT JOIN market_cap mc
            ON fr.company_id = mc.company_id
            AND fr.year = mc.year

        LEFT JOIN sectors s
            ON fr.company_id = s.company_id

        ORDER BY fr.company_id, fr.year
        """

        df = pd.read_sql(query, self.conn)

        numeric_columns = [
            "return_on_equity_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "asset_turnover",
            "free_cash_flow_cr",
            "capex_cr",
            "earnings_per_share",
            "book_value_per_share",
            "dividend_payout_ratio_pct",
            "total_debt_cr",
            "cash_from_operations_cr",
            "sales",
            "net_profit",
            "eps",
            "opm_percentage",
            "stock_price_cagr",
            "market_cap_crore",
            "enterprise_value_crore",
            "pe_ratio",
            "pb_ratio",
            "ev_ebitda",
            "dividend_yield_pct",
        ]

        for column in numeric_columns:
            if column in df.columns:
                df[column] = pd.to_numeric(
                    df[column],
                    errors="coerce",
                )

        # IMPORTANT:
        # analysis contains text such as:
        # "10 Years: 21%"
        # "5 Years: 22%"
        # "3 Years: 30%"
        # "TTM: 47%"
        #
        # Extract the actual 5-year and 10-year values.
        df["compounded_sales_growth"] = df["compounded_sales_growth"].apply(
            lambda x: self.extract_cagr(x, 5)
        )

        df["compounded_profit_growth"] = df["compounded_profit_growth"].apply(
            lambda x: self.extract_cagr(x, 5)
        )

        if "interest_coverage" in df.columns:
            df["interest_coverage"] = df["interest_coverage"].replace(
                "Debt Free",
                np.inf,
            )

        return df

    def apply_filters(self, df, filters):
        filtered_df = df.copy()

        print("\nApplying Filters")
        print("=" * 70)

        for key, value in filters.items():

            if key.endswith("_min"):
                column = key[:-4]
                operator = ">="

            elif key.endswith("_max"):
                column = key[:-4]
                operator = "<="

            else:
                continue

            if column not in filtered_df.columns:
                print(f"{column}: Column not found")
                continue

            before = len(filtered_df)

            if column == "debt_to_equity":
                non_financial = filtered_df["broad_sector"] != "Financials"

                if operator == "<=":
                    mask = (filtered_df[column] <= value) | (~non_financial)
                else:
                    mask = (filtered_df[column] >= value) | (~non_financial)

                filtered_df = filtered_df[mask]

            elif operator == ">=":
                filtered_df = filtered_df[filtered_df[column] >= value]

            else:
                filtered_df = filtered_df[filtered_df[column] <= value]

            after = len(filtered_df)

            print(f"{column:30} {before:4} -> {after:4}")

        return filtered_df

    def calculate_composite_quality_score(self, df):
        score_df = df.copy()

        metrics = [
            "return_on_equity_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "asset_turnover",
            "free_cash_flow_cr",
            "compounded_sales_growth",
            "compounded_profit_growth",
        ]

        for metric in metrics:
            if metric not in score_df.columns:
                continue

            score_df[f"{metric}_score"] = 0.0

            for sector in score_df["broad_sector"].dropna().unique():

                mask = score_df["broad_sector"] == sector
                series = score_df.loc[mask, metric]

                if series.empty:
                    continue

                p10 = series.quantile(0.10)
                p90 = series.quantile(0.90)

                clipped = series.clip(
                    lower=p10,
                    upper=p90,
                )

                if pd.isna(p10) or pd.isna(p90):
                    score_df.loc[mask, f"{metric}_score"] = 0

                elif p10 == p90:
                    score_df.loc[mask, f"{metric}_score"] = 100

                else:
                    score_df.loc[mask, f"{metric}_score"] = (
                        (clipped - p10) / (p90 - p10)
                    ) * 100

        score_df["composite_quality_score"] = (
            score_df["return_on_equity_pct_score"] * 0.25
            + score_df["net_profit_margin_pct_score"] * 0.15
            + score_df["operating_profit_margin_pct_score"] * 0.15
            + score_df["asset_turnover_score"] * 0.10
            + score_df["free_cash_flow_cr_score"] * 0.15
            + score_df["compounded_sales_growth_score"] * 0.10
            + score_df["compounded_profit_growth_score"] * 0.10
        )

        score_df["composite_quality_score"] = score_df["composite_quality_score"].round(
            2
        )

        return score_df

    def run_screener(self, preset_name):
        config = self.load_config()

        if preset_name not in config:
            raise ValueError(f"Preset '{preset_name}' not found.")

        filters = config[preset_name]

        df = self.load_data()

        filtered_df = self.apply_filters(
            df,
            filters,
        )

        filtered_df = self.calculate_composite_quality_score(filtered_df)

        filtered_df = filtered_df.sort_values(
            by="composite_quality_score",
            ascending=False,
        )

        filtered_df = filtered_df.drop_duplicates(
            subset="company_id",
            keep="first",
        )
        if preset_name == "quality_compounder":
            filtered_df = filtered_df.head(50)

        filtered_df.reset_index(
            drop=True,
            inplace=True,
        )

        filtered_df.insert(
            0,
            "Rank",
            range(1, len(filtered_df) + 1),
        )

        print("Companies:", len(filtered_df))

        return filtered_df

    def close_connection(self):
        self.conn.close()


def main():
    print("=" * 80)
    print("N100 FINANCIAL SCREENER")
    print("=" * 80)

    engine = ScreenerEngine()

    try:
        config = engine.load_config()

        print("\nAvailable Presets:")
        for preset in config:
            print(f" - {preset}")

        output_file = engine.output_path / "screener_output.xlsx"

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
        ) as writer:

            for preset_name in config:

                print(f"\nRunning Preset: {preset_name}")

                result = engine.run_screener(preset_name)

                print(f"Companies Found: {len(result)}")

                sheet_name = preset_name[:31]

                result.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )

                worksheet = writer.sheets[sheet_name]

                green_fill = PatternFill(
                    fill_type="solid",
                    start_color="C6EFCE",
                    end_color="C6EFCE",
                )

                red_fill = PatternFill(
                    fill_type="solid",
                    start_color="FFC7CE",
                    end_color="FFC7CE",
                )

                filters = config[preset_name]

                headers = {cell.value: cell.column for cell in worksheet[1]}

                for rule, threshold in filters.items():

                    if rule.endswith("_min"):
                        metric = rule[:-4]
                        comparison = "min"

                    elif rule.endswith("_max"):
                        metric = rule[:-4]
                        comparison = "max"

                    else:
                        continue

                    if metric not in headers:
                        continue

                    col = headers[metric]

                    for row in range(
                        2,
                        worksheet.max_row + 1,
                    ):

                        cell = worksheet.cell(
                            row=row,
                            column=col,
                        )

                        if cell.value is None:
                            continue

                        try:
                            value = float(cell.value)
                        except (
                            TypeError,
                            ValueError,
                        ):
                            continue

                        if comparison == "min":
                            cell.fill = green_fill if value >= threshold else red_fill
                        else:
                            cell.fill = green_fill if value <= threshold else red_fill

        print("\nExcel exported successfully.")
        print(output_file)

    finally:
        engine.close_connection()
        print("\nDatabase connection closed.")


if __name__ == "__main__":
    main()
